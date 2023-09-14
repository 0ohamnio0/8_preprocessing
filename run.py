import os
import requests
import openpyxl

# 엑셀 파일 경로와 시트 이름 설정 (본인의 파일 경로와 시트 이름으로 변경하세요)
excel_file = "C:\기초의원재보궐선거.xlsx"
sheet_name = "datatable"

# 엑셀 파일 열기
wb = openpyxl.load_workbook(excel_file)
sheet = wb[sheet_name]

# 다운로드할 파일의 저장 폴더 설정 (본인의 폴더 경로로 변경하세요)
download_folder = "C:\8회\기초의원재보궐선거_벽보"

# 현재 행 번호 추적 변수
current_row = 1  # 첫 번째 데이터 행 번호

# 링크 실행 및 파일 다운로드 및 이름 변경
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    link = row[4]  # D열에 있는 링크 (0부터 시작하는 인덱스)

    # 링크가 없는 경우 다음 행으로 이동
    if link is '':
        current_row += 1
        continue

    # 파일 다운로드
    response = requests.get(link)
    if response.status_code == 200:
        file_extension = link.split(".")[-1]  # 확장자 추출

        # 튜플의 첫 번째 요소를 문자열로 변환하고 작업 수행
        label = str(row[0])
        sanitized_label = "".join(c for c in label if c.isalnum() or c in ['-', '_', ' '])[:240]
        filename = os.path.join(download_folder, f"{sanitized_label}.pdf")
        with open(filename, 'wb') as file:
            file.write(response.content)

    # 다음 행으로 이동
    current_row += 1
